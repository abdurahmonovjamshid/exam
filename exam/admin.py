from django.contrib import admin
from .models import Exam, Question, Choice, Candidate, Attempt, Answer
from django.utils.html import format_html
from django.conf import settings
from django.urls import reverse
import openpyxl
from openpyxl.styles import PatternFill
from django.http import HttpResponse


class ChoiceInline(admin.TabularInline):
    model = Choice
    extra = 2


@admin.register(Question)
class SavolAdmin(admin.ModelAdmin):
    list_display = ("text", "exam", "qtype", "order")
    list_filter = ("exam", "qtype")
    search_fields = ("text",)
    inlines = [ChoiceInline]


@admin.register(Exam)
class ImtihonAdmin(admin.ModelAdmin):
    list_display = ("title", "duration_minutes", "shuffle_questions", "registration_link")
    search_fields = ("title",)

    def registration_link(self, obj):
        url = reverse("register_exam", args=[obj.id])
        return format_html('<a href="{}" target="_blank">Ro‘yxatga olish havolasi</a>', url)

    registration_link.short_description = "Ro‘yxatga olish havolasi"


@admin.register(Candidate)
class NomzodAdmin(admin.ModelAdmin):
    list_display = ("full_name", "phone", "region", "work_position", "hr_manager", "created_at")
    search_fields = ("full_name", "region", "work_position")


def strip_tz(dt):
    """Excel uchun timezone olib tashlash"""
    return dt.replace(tzinfo=None) if dt else None


@admin.register(Attempt)
class UrinishAdmin(admin.ModelAdmin):
    list_display = (
        "candidate",
        "exam",
        "score",
        "started_at",
        "submitted_at",
        "exam_link",
    )
    list_filter = ("exam", "submitted_at")
    search_fields = ("candidate__full_name", "exam__title", "token")
    readonly_fields = ("token", "exam_link", "started_at", "submitted_at")

    actions = ["export_attempts_excel"]

    def exam_link(self, obj):
        if not obj.token:
            return "-"
        base_url = getattr(settings, "SITE_URL", "http://127.0.0.1:8000")
        url = f"{base_url}/exam/{obj.token}/"
        return format_html('<a href="{}" target="_blank">Imtihon havolasi</a>', url)

    exam_link.short_description = "Imtihon havolasi"

    def export_attempts_excel(self, request, queryset):
        """
        Tanlangan urinishlarni Excelga eksport qilish
        - Nomzod ma’lumoti
        - Imtihon ma’lumoti
        - Ball
        - Har bir savol ustun ko‘rinishida
        - Javob matni: to‘g‘ri yashil, noto‘g‘ri qizil
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Urinishlar"

        # Savollar to‘plami
        all_questions = []
        for attempt in queryset:
            for q in attempt.exam.questions.all():
                if q not in all_questions:
                    all_questions.append(q)

        # Sarlavha qatori
        headers = [
            "Nomzod",
            "Telefon",
            "Hudud",
            "Ish o‘rni",
            "HR menejer",
            "Imtihon",
            "Ball",
            "Boshlangan vaqt",
            "Yakunlangan vaqt",
        ]
        headers += [q.text[:50] for q in all_questions]  # savol matnini qisqartirish
        ws.append(headers)

        # Ranglar
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Ma’lumotlar qatori
        for attempt in queryset:
            row = [
                attempt.candidate.full_name,
                attempt.candidate.phone,
                attempt.candidate.region,
                attempt.candidate.work_position,
                attempt.candidate.hr_manager,
                attempt.exam.title,
                attempt.score,
                strip_tz(attempt.started_at),
                strip_tz(attempt.submitted_at),
            ]

            answers_map = {a.question_id: a for a in attempt.answers.all()}

            for q in all_questions:
                ans = answers_map.get(q.id)
                if ans:
                    if ans.choice:
                        cell_value = ans.choice.text
                    else:
                        cell_value = ans.text_answer
                else:
                    cell_value = ""
                row.append(cell_value)

            ws.append(row)

            # Rang berish
            excel_row = ws.max_row
            for idx, q in enumerate(all_questions, start=len(headers) - len(all_questions) + 1):
                ans = answers_map.get(q.id)
                cell = ws.cell(row=excel_row, column=idx + 1)
                if ans:
                    if ans.is_correct:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

        # Sana va vaqt formatlash
        for col in ("H", "I"):  # Boshlangan va Yakunlangan
            for cell in ws[col]:
                if cell.row == 1:
                    continue
                if cell.value:
                    cell.number_format = "DD.MM.YYYY HH:MM"

        # Javob
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="urinishlar.xlsx"'
        wb.save(response)
        return response

    export_attempts_excel.short_description = "Tanlanganlarni Excelga yuklash"


@admin.register(Answer)
class JavobAdmin(admin.ModelAdmin):
    list_display = ("attempt", "question", "choice", "text_answer", "is_correct")
    list_filter = ("is_correct", "question__exam")
    search_fields = ("attempt__candidate__full_name", "question__text")
